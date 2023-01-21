from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.common.by import By
import datetime
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import simpledialog
import pyautogui


def formatar_num(num):
    if str(num).startswith('$'):
        num = str(num)[1:]
    if str(num).count("B") > 0:
        num = str(num).replace('B', '')
        num = float(num) * 1000000000
    if str(num).count("M") > 0:
        num = str(num).replace('M', '')
        num = float(num) * 1000000
    if str(num).count("K") > 0:
        num = str(num).replace('K', '')
        num = float(num) * 1000
    if str(num).count(".") > 0:
        num = str(num).replace(".", ",")
    return num

# definir data de hoje
agora = datetime.datetime.now()
data = agora.strftime('%d/%m/%y')

#caixa input datas
root = tk.Tk()
root.withdraw()

# Recebe data FROM
validacao_from = False
while validacao_from == False:
    input_from = simpledialog.askstring("Data From", "Digite a data FROM no formato yyyy-mm-dd")
    if str(input_from).count("-") == 2:
        input_from_validacao = str(input_from).split("-")
        if len(input_from_validacao) != 3:
            validacao_from = False
        else:
            if len(input_from_validacao[0]) == 4 and len(input_from_validacao[1]) == 2 and len(input_from_validacao[2]) == 2:
                validacao_from = True


# Recebe data UNTIL
validacao_until = False
while validacao_until == False:
    input_until = simpledialog.askstring("Data Until", "Digite a data UNTIL no formato yyyy-mm-dd")
    if str(input_until).count("-") == 2:
        input_until_validacao = str(input_until).split("-")
        if len(input_until_validacao) != 3:
            validacao_until = False
        else:
            if len(input_until_validacao[0]) == 4 and len(input_until_validacao[1]) == 2 and len(input_until_validacao[2]) == 2:
                validacao_until = True

root.destroy()


#ocultar navegador do usuário
opcao = webdriver.ChromeOptions()
opcao.add_argument("--headless")
#criar navegador inflation
navegador = webdriver.Chrome(options=opcao)

# acessar site
navegador.get('https://tradingeconomics.com/calendar/inflation')
navegador.implicitly_wait(5)

# country
navegador.find_element(By.XPATH, '//*[@id="aspnetForm"]/div[4]/div/div/table/tbody/tr/td[1]/div/button/span').click()
# clear
navegador.find_element(By.XPATH, '//*[@id="te-c-main-countries"]/div/div[2]/div[1]/a').click()
# selecionar EUA
navegador.find_element(By.XPATH, '//*[@id="te-c-all"]/ul[4]/li[27]/input').click()
sleep(1)
# save
element = navegador.find_element(By.XPATH, "//*[@id='te-c-main-countries']/div/div[2]/div[3]/a")
actions = ActionChains(navegador)  # Move o cursor do mouse até o elemento
actions.move_to_element(element).perform()
element.click()  # Clica no elemento
#clicar filtro data
navegador.find_element(By.XPATH, '//*[@id="aspnetForm"]/div[4]/div/div/table/tbody/tr/td[1]/div/div[1]/button').click()
#clicar custom
navegador.find_element(By.XPATH, '//*[@id="aspnetForm"]/div[4]/div/div/table/tbody/tr/td[1]/div/div[1]/ul/li[12]/a').click()


#DATA FROM - apagar campo e preencher com data escolhida
from_data = navegador.find_element(By.XPATH, '//*[@id="startDate"]') #cria elemento
from_data.clear() #apaga texto
from_data.send_keys(input_from) #escreve data

#DATA UNTIL - apagar campo e preencher com data escolhida
until_data = navegador.find_element(By.XPATH, '//*[@id="endDate"]') #cria elemento
until_data.clear() #apaga texto
until_data.send_keys(input_until) #escreve data

#clicar em submit
navegador.find_element(By.XPATH, '//*[@id="datesDiv"]/div/span[3]/button').click()


# extrair dados inflation
titulos = navegador.find_elements(By.CLASS_NAME, 'calendar-event')
actual = navegador.find_elements(By.ID, 'actual')
consensus = navegador.find_elements(By.ID, 'consensus')
data_referencia = navegador.find_elements(By.CSS_SELECTOR, 'span.calendar-reference')


#criar navegador labour
navegador_labour = webdriver.Chrome(options=opcao)

# acessar site
navegador_labour.get('https://tradingeconomics.com/calendar/labour')
navegador_labour.implicitly_wait(5)

# country
navegador_labour.find_element(By.XPATH, '//*[@id="aspnetForm"]/div[4]/div/div/table/tbody/tr/td[1]/div/button/span').click()
# clear
navegador_labour.find_element(By.XPATH, '//*[@id="te-c-main-countries"]/div/div[2]/div[1]/a').click()
# selecionar EUA
navegador_labour.find_element(By.XPATH, '//*[@id="te-c-all"]/ul[4]/li[27]/input').click()
sleep(1)
# save
element = navegador_labour.find_element(By.XPATH, "//*[@id='te-c-main-countries']/div/div[2]/div[3]/a")
actions = ActionChains(navegador_labour)  # Move o cursor do mouse até o elemento
actions.move_to_element(element).perform()
element.click()  # Clica no elemento
#clicar filtro data
navegador_labour.find_element(By.XPATH, '//*[@id="aspnetForm"]/div[4]/div/div/table/tbody/tr/td[1]/div/div[1]/button').click()
#clicar custom
navegador_labour.find_element(By.XPATH, '//*[@id="aspnetForm"]/div[4]/div/div/table/tbody/tr/td[1]/div/div[1]/ul/li[12]/a').click()

#DATA FROM - apagar campo e preencher com data escolhida
from_data = navegador_labour.find_element(By.XPATH, '//*[@id="startDate"]') #cria elemento
from_data.clear() #apaga texto
from_data.send_keys(input_from) #escreve data

#DATA UNTIL - apagar campo e preencher com data escolhida
until_data = navegador_labour.find_element(By.XPATH, '//*[@id="endDate"]') #cria elemento
until_data.clear() #apaga texto
until_data.send_keys(input_until) #escreve data

#clicar em submit
navegador_labour.find_element(By.XPATH, '//*[@id="datesDiv"]/div/span[3]/button').click()

#extrair dados labour
titulos.extend(navegador_labour.find_elements(By.CSS_SELECTOR, 'a.calendar-event'))
actual.extend(navegador_labour.find_elements(By.ID, 'actual'))
consensus.extend(navegador_labour.find_elements(By.ID, 'consensus'))
data_referencia.extend(navegador_labour.find_elements(By.CSS_SELECTOR, 'span.calendar-reference'))


# listas de dados
indice_individual = []  # individual
indice_todos = []  # todos

indicadores = ['PCE Price Index MoM',
                'PCE Price Index YoY',
                'Core PCE Price Index MoM',
                'Core PCE Price Index YoY',
                'PPI MoM',
                'PPI YoY',
                'Core PPI MoM',
                'Core PPI YoY',
                'Inflation Rate MoM',
                'Inflation Rate YoY',
                'Core Inflation Rate MoM',
                'Core Inflation Rate YoY',
                'Non Farm Payrolls',
                'Unemployment Rate',
                'Average Hourly Earnings MoM',
                'Average Hourly Earnings YoY',
                'Initial Jobless Claims']

#cria lista com dados necessários
for c in range(0, len(titulos)):
    if str(titulos[c].text) in indicadores:
        indice_individual.append(titulos[c].text)
        indice_individual.append(formatar_num(actual[c].text))
        indice_individual.append(formatar_num(consensus[c].text))
        indice_individual.append(data_referencia[c].text)
        indice_todos.append(indice_individual)
        indice_individual = []

#liberar memoria
titulos = []
actual = []
data_referencia = []

#carrega planilha para ser editada
planilha_path = r'write xlsx path'
wb = load_workbook(planilha_path, keep_vba=True)
ws = wb.worksheets[2]

for linha in range(0, len(indice_todos)):
    for coluna in range(0, 4):
        if indice_todos[linha][coluna] != "":
            ws.cell(row=linha+2, column=coluna+1).value = indice_todos[linha][coluna]
        else:
            ws.cell(row=linha + 2, column=coluna + 1).value = 0
    ws.cell(row=linha + 2, column=5).value = data


wb.save(planilha_path)


# fecha os navegadores
navegador.quit()
navegador_labour.quit()


pyautogui.alert('Planilha atualizada.', button='OK', title='Atualizada')

os.startfile(planilha_path)
